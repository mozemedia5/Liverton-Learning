import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# ==================== WORD PROCESSORS SHEET ====================
ws1 = wb.active
ws1.title = "Word Processors"

# Header styling
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
category_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
category_font = Font(bold=True, color='FFFFFF', size=11)

# Title
ws1['A1'] = 'WORD PROCESSORS (Word / Docs)'
ws1['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws1.merge_cells('A1:C1')

# Headers
headers1 = ['Category', 'Tool', 'Description']
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=2, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data for Word Processors
word_data = [
    ('Text Manipulation', 'Cut', 'Remove selected text and place it on clipboard'),
    ('', 'Copy', 'Duplicate selected text to clipboard'),
    ('', 'Paste', 'Insert clipboard content at cursor position'),
    ('', 'Paste Special', 'Insert with specific formatting options'),
    ('', 'Undo', 'Reverse the last action'),
    ('', 'Redo', 'Reapply the last undone action'),
    ('', 'Select All', 'Highlight all content in the document'),
    ('', 'Find', 'Search for specific text'),
    ('', 'Find & Replace', 'Search and substitute text throughout document'),
    ('Review & Proofing', 'Spell Check', 'Identify and correct spelling errors'),
    ('', 'Grammar Check', 'Detect grammar and style issues'),
    ('', 'Word Count', 'Display statistics (words, characters, pages)'),
    ('', 'Translate', 'Convert text to different languages'),
    ('', 'Track Changes', 'Record edits for review'),
    ('', 'Suggesting Mode', 'Propose changes without altering original'),
    ('', 'Comments', 'Add notes and feedback'),
    ('', 'Review Changes', 'Accept or reject tracked modifications'),
    ('', 'Compare Documents', 'Highlight differences between versions'),
    ('Voice & Input', 'Voice Typing', 'Dictate text using speech recognition'),
]

current_category = ''
for row_idx, (category, tool, desc) in enumerate(word_data, 3):
    ws1.cell(row=row_idx, column=1, value=category if category else current_category)
    ws1.cell(row=row_idx, column=2, value=tool)
    ws1.cell(row=row_idx, column=3, value=desc)
    if category:
        current_category = category
        ws1.cell(row=row_idx, column=1).fill = category_fill
        ws1.cell(row=row_idx, column=1).font = category_font

# ==================== SPREADSHEETS SHEET ====================
ws2 = wb.create_sheet(title="Spreadsheets")

ws2['A1'] = 'SPREADSHEETS (Excel / Sheets)'
ws2['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws2.merge_cells('A1:C1')

for col, header in enumerate(headers1, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

spreadsheet_data = [
    ('Cell Editing', 'Insert Row/Column', 'Add new rows or columns'),
    ('', 'Delete Row/Column', 'Remove rows or columns'),
    ('', 'Merge Cells', 'Combine multiple cells into one'),
    ('', 'Split Cells', 'Divide merged cells'),
    ('', 'Wrap Text', 'Allow text to display on multiple lines'),
    ('', 'Text Rotation', 'Change text orientation'),
    ('', 'Clear Contents', 'Remove data while keeping formatting'),
    ('', 'Fill Down', 'Copy formula/content to cells below'),
    ('', 'Fill Right', 'Copy formula/content to cells right'),
    ('', 'Find & Replace', 'Search and replace cell content'),
    ('Data Manipulation', 'Sort', 'Arrange data alphabetically or numerically'),
    ('', 'Filter', 'Display only rows meeting criteria'),
    ('', 'Remove Duplicates', 'Eliminate repeated entries'),
    ('', 'Data Validation', 'Restrict input with rules'),
    ('', 'Freeze Rows/Columns', 'Keep headers visible while scrolling'),
    ('', 'Hide / Unhide', 'Conceal or reveal rows/columns'),
    ('', 'Group / Ungroup', 'Create collapsible data sections'),
    ('Calculation', 'Functions', 'Built-in formulas (SUM, AVERAGE, IF, etc.)'),
    ('', 'AutoSum', 'Quick sum of selected cells'),
    ('', 'Formula Bar', 'Edit cell formulas'),
    ('', 'Recalculate', 'Update all formula results'),
]

current_category = ''
for row_idx, (category, tool, desc) in enumerate(spreadsheet_data, 3):
    ws2.cell(row=row_idx, column=1, value=category if category else current_category)
    ws2.cell(row=row_idx, column=2, value=tool)
    ws2.cell(row=row_idx, column=3, value=desc)
    if category:
        current_category = category
        ws2.cell(row=row_idx, column=1).fill = category_fill
        ws2.cell(row=row_idx, column=1).font = category_font

# ==================== PRESENTATIONS SHEET ====================
ws3 = wb.create_sheet(title="Presentations")

ws3['A1'] = 'PRESENTATIONS (PowerPoint / Slides)'
ws3['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws3.merge_cells('A1:C1')

for col, header in enumerate(headers1, 1):
    cell = ws3.cell(row=2, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

presentation_data = [
    ('Slide Editing', 'New Slide', 'Add a blank slide'),
    ('', 'Duplicate Slide', 'Create a copy of existing slide'),
    ('', 'Delete Slide', 'Remove a slide'),
    ('', 'Reorder Slides', 'Change slide sequence'),
    ('', 'Slide Layouts', 'Apply pre-designed arrangements'),
    ('', 'Reset Slide', 'Restore layout to default'),
    ('Object Editing', 'Align Objects', 'Position multiple items (left, center, right)'),
    ('', 'Arrange Layers', 'Bring to front / Send to back'),
    ('', 'Group Objects', 'Combine multiple items'),
    ('', 'Ungroup', 'Separate grouped items'),
    ('', 'Rotate', 'Turn objects at angles'),
    ('', 'Flip', 'Mirror objects horizontally or vertically'),
    ('', 'Crop', 'Trim image edges'),
    ('', 'Resize', 'Change object dimensions'),
    ('Visual Adjustments', 'Backgrounds', 'Change slide background color/image'),
    ('', 'Themes', 'Apply design templates'),
    ('', 'Master Slide Edits', 'Modify template for all slides'),
    ('Animation & Timing', 'Entrance Animations', 'Effects for objects appearing'),
    ('', 'Exit Animations', 'Effects for objects disappearing'),
    ('', 'Motion Paths', 'Define movement trajectories'),
    ('', 'Timing Control', 'Adjust animation speed and delays'),
    ('', 'Trigger Actions', 'Set animation start conditions'),
]

current_category = ''
for row_idx, (category, tool, desc) in enumerate(presentation_data, 3):
    ws3.cell(row=row_idx, column=1, value=category if category else current_category)
    ws3.cell(row=row_idx, column=2, value=tool)
    ws3.cell(row=row_idx, column=3, value=desc)
    if category:
        current_category = category
        ws3.cell(row=row_idx, column=1).fill = category_fill
        ws3.cell(row=row_idx, column=1).font = category_font

# ==================== QUICK REFERENCE SHEET ====================
ws4 = wb.create_sheet(title="Quick Reference")

ws4['A1'] = 'QUICK REFERENCE SUMMARY'
ws4['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws4.merge_cells('A1:D1')

headers_ref = ['Category', 'Primary Tools', 'Advanced Features', 'Total Tools']
for col, header in enumerate(headers_ref, 1):
    cell = ws4.cell(row=2, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

ref_data = [
    ('Word Processors', 'Cut, Copy, Paste, Find & Replace', 'Track Changes, Compare, Voice Typing', '19'),
    ('Spreadsheets', 'Insert/Delete, Sort, Filter', 'Data Validation, Functions, Freeze', '21'),
    ('Presentations', 'New/Duplicate Slides, Layouts', 'Animations, Master Slides, Grouping', '22'),
]

for row_idx, (cat, primary, advanced, total) in enumerate(ref_data, 3):
    ws4.cell(row=row_idx, column=1, value=cat)
    ws4.cell(row=row_idx, column=2, value=primary)
    ws4.cell(row=row_idx, column=3, value=advanced)
    ws4.cell(row=row_idx, column=4, value=total)

# Adjust column widths for all sheets
for ws in [ws1, ws2, ws3]:
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 35
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 15

# Save workbook
wb.save('/workspace/liverton-learning/Editing_Tools_Spreadsheet.xlsx')
print("Created: Editing_Tools_Spreadsheet.xlsx")
